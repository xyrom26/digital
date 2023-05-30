#Libraries
import keyboard
import pyautogui
import time
import sys
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime       
import pandas as pd



options = Options()
options.add_experimental_option("detach", True)


name = sys.argv[1]
email = sys.argv[2]
data = sys.argv[3]


print("We will start at 3s")


wb = load_workbook(data)
ws = wb.active


row_count = sum(1 for row in ws if any(cell.value is not None for cell in row)) 
currRow = 2

#Install the chrome driver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
driver.maximize_window()

#Link to be open
driver.get("https://www.digitalocean.com/company/contact/abuse#intrusion")
time.sleep(3)
agree = driver.find_element("xpath",'.//*[@id="truste-consent-button"]')
agree.click()
pyautogui.hotkey('ctrl', '-')
pyautogui.hotkey('ctrl', '-')
pyautogui.hotkey('ctrl', '-')
pyautogui.hotkey('ctrl', '-')
pyautogui.hotkey('ctrl', '-')
pyautogui.hotkey('ctrl', '-')
pyautogui.hotkey('ctrl', '-')


# Set the initial value for last_characteristic to None
last_characteristic = None

row_number = 0

for row in ws.iter_rows(min_row=2, max_row=row_count, values_only=True):
    # Wait for the page to fully load
    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_element_located((By.XPATH, "//body")))

    # Check if the page is fully loaded using JavaScript
    is_loaded = driver.execute_script("return document.readyState")

    if is_loaded == "complete":
        print("Page is fully loaded")

        # avoid duplicate report on ip
        if row[1] == last_characteristic:
            continue  # skip processing this row if it has the same characteristic as the previous row
        else:
            last_characteristic = row[1]  # update the last characteristic variable
            logs = [str(cell.value) if cell.value is not None else '' for cell in ws[currRow][:15]] # columns A to O
            
            driver.refresh()
            driver.execute_script("window.scrollTo(0, 0)")

            # Scan/WebScan
            if row[15] == "Scans":
                abuseType = row[15]

              
                atype = driver.find_element(By.ID, 'react-select-3-input')
                atype.click()
                atype = driver.find_element("xpath", './/*[@id="react-select-3-option-1"]')
                atype.click()

                #Check Abusetype
                if row[20] == "WebScan":

                    currRow +=1

                    # WebScan type of Scan
                    abtclick = driver.find_element("xpath",'.//*[@id="intrusion_sub_category-webscan"]')
                    abtclick.click()

                    # Name
                    uname = driver.find_element("xpath",'.//*[@id="reporter_name"]')
                    uname.send_keys(name)
                    

                    # Email
                    email_sender = driver.find_element("xpath",'.//*[@id="email"]')
                    email_sender.send_keys(email)
           

                    # Abuse Evidence logs
                    evidence = driver.find_element("xpath",'.//*[@id="abuse_evidence"]')
                    evidence.send_keys(logs) # This coce  is a generator expression that converts each non-None value in the row list to a string, then joins them together with a space separator.
           

                    # Source IP
                    scr_ip = driver.find_element("xpath",'.//*[@id="source_ip"]')
                    scr_ip.send_keys(row[1])
              

                    # Date
                    date = driver.find_element("xpath",'.//*[@id="date"]')
                    date_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    date_obj = datetime.strptime(date_str, "%b %d, %Y, %I:%M:%S %p") # This code converts a date string in the format of "MMM DD, YYYY, HH:MM:SS AM/PM" into a datetime object in Python.
                    date_input = driver.find_element("xpath", './/*[@id="date"]')
                    date_input.send_keys(date_obj.strftime("%m-%d-%y"))

                    time = driver.find_element("xpath",'.//*[@id="time"]')
                    time_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    time_obj = datetime.strptime(time_str, "%b %d, %Y, %I:%M:%S %p")
                    time_input = driver.find_element("xpath", './/*[@id="time"]')
                    time_input.send_keys(time_obj.strftime("%H:%m"))

                    # Abuse time zone
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-input"]')
                    time_zone.click()
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-option-28"]')
                    time_zone.click()

                    # disktination port/protocol
                    disk_port = row[6]
                    time_input = driver.find_element("xpath", './/*[@id="destination_port"]')
                    time_input.send_keys(disk_port)

                    # destination ip addresses:
                    disk_ip = row[5]
                    time_input = driver.find_element("xpath", './/*[@id="destination_ip_addresses"]')
                    time_input.send_keys(disk_ip)

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()


                    # Submit form
                    submit_report = driver.find_element("xpath",'./html/body/div[1]/div[3]/div/div[1]/section/div/div/div/form/button')
                    submit_report.click()

                    print(logs)

                    row_number += 1

                    # Scroll up to the top of the web page
                    driver.execute_script("window.scrollTo(0, 0)")
                    driver.refresh()
                    driver.execute_script("window.scrollTo(0, 0)")

                
                elif row[20] == "PortScan":

                    currRow +=1

                    # PortScan type of Scan
                    abtclick = driver.find_element("xpath",'.//*[@id="intrusion_sub_category-portscan"]')
                    abtclick.click()
                   

                    # Name
                    uname = driver.find_element("xpath",'.//*[@id="reporter_name"]')
                    uname.send_keys(name)
                    

                    # Email
                    email_sender = driver.find_element("xpath",'.//*[@id="email"]')
                    email_sender.send_keys(email)
           

                    # Abuse Evidence logs
                    evidence = driver.find_element("xpath",'.//*[@id="abuse_evidence"]')
                    evidence.send_keys(logs) # This coce  is a generator expression that converts each non-None value in the row list to a string, then joins them together with a space separator.
           

                    # Source IP
                    scr_ip = driver.find_element("xpath",'.//*[@id="source_ip"]')
                    scr_ip.send_keys(row[1])
              

                    # Date
                    date = driver.find_element("xpath",'.//*[@id="date"]')
                    date_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    date_obj = datetime.strptime(date_str, "%b %d, %Y, %I:%M:%S %p") # This code converts a date string in the format of "MMM DD, YYYY, HH:MM:SS AM/PM" into a datetime object in Python.
                    date_input = driver.find_element("xpath", './/*[@id="date"]')
                    date_input.send_keys(date_obj.strftime("%m-%d-%y"))

                    time = driver.find_element("xpath",'.//*[@id="time"]')
                    time_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    time_obj = datetime.strptime(time_str, "%b %d, %Y, %I:%M:%S %p")
                    time_input = driver.find_element("xpath", './/*[@id="time"]')
                    time_input.send_keys(time_obj.strftime("%H:%m"))

                    # Abuse time zone
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-input"]')
                    time_zone.click()
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-option-28"]')
                    time_zone.click()

                    # disktination port/protocol
                    disk_port = row[6]
                    time_input = driver.find_element("xpath", './/*[@id="destination_port"]')
                    time_input.send_keys(disk_port)

                    # destination ip addresses:
                    disk_ip = row[5]
                    time_input = driver.find_element("xpath", './/*[@id="destination_ip_addresses"]')
                    time_input.send_keys(disk_ip)

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()


                    # Submit form
                    submit_report = driver.find_element("xpath",'./html/body/div[1]/div[3]/div/div[1]/section/div/div/div/form/button')
                    submit_report.click()
                    

                    print(logs)

                    row_number += 1

                    # Scroll up to the top of the web page
                    driver.execute_script("window.scrollTo(0, 0)")
                    driver.refresh()
                    driver.execute_script("window.scrollTo(0, 0)")

                else:
                  
                    currRow +=1

                    # Others type of Scan
                    abtclick = driver.find_element("xpath",'.//*[@id="intrusion_sub_category-others"]')
                    abtclick.click()
                    keyboard.press_and_release('tab')

                    # Name
                    uname = driver.find_element("xpath",'.//*[@id="reporter_name"]')
                    uname.send_keys(name)
                    

                    # Email
                    email_sender = driver.find_element("xpath",'.//*[@id="email"]')
                    email_sender.send_keys(email)
           

                    # Abuse Evidence logs
                    evidence = driver.find_element("xpath",'.//*[@id="abuse_evidence"]')
                    evidence.send_keys(logs) # This coce  is a generator expression that converts each non-None value in the row list to a string, then joins them together with a space separator.
           

                    # Source IP
                    scr_ip = driver.find_element("xpath",'.//*[@id="source_ip"]')
                    scr_ip.send_keys(row[1])
              

                    # Date
                    date = driver.find_element("xpath",'.//*[@id="date"]')
                    date_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    date_obj = datetime.strptime(date_str, "%b %d, %Y, %I:%M:%S %p") # This code converts a date string in the format of "MMM DD, YYYY, HH:MM:SS AM/PM" into a datetime object in Python.
                    date_input = driver.find_element("xpath", './/*[@id="date"]')
                    date_input.send_keys(date_obj.strftime("%m-%d-%y"))

                    time = driver.find_element("xpath",'.//*[@id="time"]')
                    time_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    time_obj = datetime.strptime(time_str, "%b %d, %Y, %I:%M:%S %p")
                    time_input = driver.find_element("xpath", './/*[@id="time"]')
                    time_input.send_keys(time_obj.strftime("%H:%m"))

                    # Abuse time zone
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-input"]')
                    time_zone.click()
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-option-28"]')
                    time_zone.click()

                    # disktination port/protocol
                    disk_port = row[6]
                    time_input = driver.find_element("xpath", './/*[@id="destination_port"]')
                    time_input.send_keys(disk_port)

                    # destination ip addresses:
                    disk_ip = row[5]
                    time_input = driver.find_element("xpath", './/*[@id="destination_ip_addresses"]')
                    time_input.send_keys(disk_ip)

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()


                    # Submit form
                    submit_report = driver.find_element("xpath",'./html/body/div[1]/div[3]/div/div[1]/section/div/div/div/form/button')
                    submit_report.click()

                    print(logs)

                    row_number += 1
                
                    # Scroll up to the top of the web page
                    driver.execute_script("window.scrollTo(0, 0)")
                    driver.refresh()
                    driver.execute_script("window.scrollTo(0, 0)")
            # Bruteforce abtype   
            elif row[15] == "Bruteforce":
                abuseType = row[15]

                #ScanType
                atype = driver.find_element(By.ID, 'react-select-3-input')
                atype.click()
                atype = driver.find_element("xpath", './/*[@id="react-select-3-option-0"]')
                atype.click()

                #AbuseTye SSH
                if str(row[16]) == "True" or str(row[16]) == "TRUE":
                    
                    currRow +=1

                    abtclick = driver.find_element("xpath",'.//*[@id="intrusion_sub_category-ssh"]')
                    abtclick.click()

                    # Name
                    uname = driver.find_element("xpath",'.//*[@id="reporter_name"]')
                    uname.send_keys(name)
                    

                    # Email
                    email_sender = driver.find_element("xpath",'.//*[@id="email"]')
                    email_sender.send_keys(email)
           

                    # Abuse Evidence logs
                    evidence = driver.find_element("xpath",'.//*[@id="abuse_evidence"]')
                    evidence.send_keys(logs) # This coce  is a generator expression that converts each non-None value in the row list to a string, then joins them together with a space separator.
           

                    # Source IP
                    scr_ip = driver.find_element("xpath",'.//*[@id="source_ip"]')
                    scr_ip.send_keys(row[1])
              

                    # Date
                    date = driver.find_element("xpath",'.//*[@id="date"]')
                    date_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    date_obj = datetime.strptime(date_str, "%b %d, %Y, %I:%M:%S %p") # This code converts a date string in the format of "MMM DD, YYYY, HH:MM:SS AM/PM" into a datetime object in Python.
                    date_input = driver.find_element("xpath", './/*[@id="date"]')
                    date_input.send_keys(date_obj.strftime("%m-%d-%y"))

                    time = driver.find_element("xpath",'.//*[@id="time"]')
                    time_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    time_obj = datetime.strptime(time_str, "%b %d, %Y, %I:%M:%S %p")
                    time_input = driver.find_element("xpath", './/*[@id="time"]')
                    time_input.send_keys(time_obj.strftime("%H:%m"))

                    # Abuse time zone
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-input"]')
                    time_zone.click()
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-option-28"]')
                    time_zone.click()

                    # disktination port/protocol
                    disk_port = row[6]
                    time_input = driver.find_element("xpath", './/*[@id="destination_port"]')
                    time_input.send_keys(disk_port)

                    # destination ip addresses:
                    disk_ip = row[5]
                    time_input = driver.find_element("xpath", './/*[@id="destination_ip_addresses"]')
                    time_input.send_keys(disk_ip)

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()


                    # Submit form
                    submit_report = driver.find_element("xpath",'./html/body/div[1]/div[3]/div/div[1]/section/div/div/div/form/button')
                    submit_report.click()

                    print(logs)

                    row_number += 1
                
                    # Scroll up to the top of the web page
                    driver.execute_script("window.scrollTo(0, 0)")
                    driver.refresh()
                    driver.execute_script("window.scrollTo(0, 0)")

                #AbuseTye WORDPRESS
                elif str(row[17]) == "True" or str(row[17]) == "TRUE":
                    
                    currRow +=1

                    abtclick = driver.find_element("xpath",'.//*[@id="intrusion_sub_category-wordpress"]')
                    abtclick.click()

                    # Name
                    uname = driver.find_element("xpath",'.//*[@id="reporter_name"]')
                    uname.send_keys(name)
                    

                    # Email
                    email_sender = driver.find_element("xpath",'.//*[@id="email"]')
                    email_sender.send_keys(email)
           

                    # Abuse Evidence logs
                    evidence = driver.find_element("xpath",'.//*[@id="abuse_evidence"]')
                    evidence.send_keys(logs) # This coce  is a generator expression that converts each non-None value in the row list to a string, then joins them together with a space separator.
           

                    # Source IP
                    scr_ip = driver.find_element("xpath",'.//*[@id="source_ip"]')
                    scr_ip.send_keys(row[1])
              

                    # Date
                    date = driver.find_element("xpath",'.//*[@id="date"]')
                    date_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    date_obj = datetime.strptime(date_str, "%b %d, %Y, %I:%M:%S %p") # This code converts a date string in the format of "MMM DD, YYYY, HH:MM:SS AM/PM" into a datetime object in Python.
                    date_input = driver.find_element("xpath", './/*[@id="date"]')
                    date_input.send_keys(date_obj.strftime("%m-%d-%y"))

                    time = driver.find_element("xpath",'.//*[@id="time"]')
                    time_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    time_obj = datetime.strptime(time_str, "%b %d, %Y, %I:%M:%S %p")
                    time_input = driver.find_element("xpath", './/*[@id="time"]')
                    time_input.send_keys(time_obj.strftime("%H:%m"))

                    # Abuse time zone
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-input"]')
                    time_zone.click()
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-option-28"]')
                    time_zone.click()

                    # disktination port/protocol
                    disk_port = row[6]
                    time_input = driver.find_element("xpath", './/*[@id="destination_port"]')
                    time_input.send_keys(disk_port)

                    # destination ip addresses:
                    disk_ip = row[5]
                    time_input = driver.find_element("xpath", './/*[@id="destination_ip_addresses"]')
                    time_input.send_keys(disk_ip)

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()


                    # Submit form
                    submit_report = driver.find_element("xpath",'./html/body/div[1]/div[3]/div/div[1]/section/div/div/div/form/button')
                    submit_report.click()

                    print(logs)

                    row_number += 1
                
                    # Scroll up to the top of the web page
                    driver.execute_script("window.scrollTo(0, 0)")
                    driver.refresh()
                    driver.execute_script("window.scrollTo(0, 0)")
                
                #AbuseTye smtp/imap
                elif str(row[18]) == "True" or str(row[18]) == "TRUE":
                    
                    currRow +=1

                    abtclick = driver.find_element("xpath",'.//*[@id="intrusion_sub_category-smtp/imap"]')
                    abtclick.click()

                    # Name
                    uname = driver.find_element("xpath",'.//*[@id="reporter_name"]')
                    uname.send_keys(name)
                    

                    # Email
                    email_sender = driver.find_element("xpath",'.//*[@id="email"]')
                    email_sender.send_keys(email)
           

                    # Abuse Evidence logs
                    evidence = driver.find_element("xpath",'.//*[@id="abuse_evidence"]')
                    evidence.send_keys(logs) # This coce  is a generator expression that converts each non-None value in the row list to a string, then joins them together with a space separator.
           

                    # Source IP
                    scr_ip = driver.find_element("xpath",'.//*[@id="source_ip"]')
                    scr_ip.send_keys(row[1])
              

                    # Date
                    date = driver.find_element("xpath",'.//*[@id="date"]')
                    date_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    date_obj = datetime.strptime(date_str, "%b %d, %Y, %I:%M:%S %p") # This code converts a date string in the format of "MMM DD, YYYY, HH:MM:SS AM/PM" into a datetime object in Python.
                    date_input = driver.find_element("xpath", './/*[@id="date"]')
                    date_input.send_keys(date_obj.strftime("%m-%d-%y"))

                    time = driver.find_element("xpath",'.//*[@id="time"]')
                    time_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    time_obj = datetime.strptime(time_str, "%b %d, %Y, %I:%M:%S %p")
                    time_input = driver.find_element("xpath", './/*[@id="time"]')
                    time_input.send_keys(time_obj.strftime("%H:%m"))

                    # Abuse time zone
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-input"]')
                    time_zone.click()
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-option-28"]')
                    time_zone.click()

                    # disktination port/protocol
                    disk_port = row[6]
                    time_input = driver.find_element("xpath", './/*[@id="destination_port"]')
                    time_input.send_keys(disk_port)

                    # destination ip addresses:
                    disk_ip = row[5]
                    time_input = driver.find_element("xpath", './/*[@id="destination_ip_addresses"]')
                    time_input.send_keys(disk_ip)

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()


                    # Submit form
                    submit_report = driver.find_element("xpath",'./html/body/div[1]/div[3]/div/div[1]/section/div/div/div/form/button')
                    submit_report.click()

                    print(logs)

                    row_number += 1
                
                    # Scroll up to the top of the web page
                    driver.execute_script("window.scrollTo(0, 0)")
                    driver.refresh()
                    driver.execute_script("window.scrollTo(0, 0)")

                #AbuseTye sip
                elif str(row[19]) == "True" or str(row[19]) == "TRUE":
                    
                    currRow +=1

                    abtclick = driver.find_element("xpath",'.//*[@id="intrusion_sub_category-sip"]')
                    abtclick.click()

                    # Name
                    uname = driver.find_element("xpath",'.//*[@id="reporter_name"]')
                    uname.send_keys(name)
                    

                    # Email
                    email_sender = driver.find_element("xpath",'.//*[@id="email"]')
                    email_sender.send_keys(email)
           

                    # Abuse Evidence logs
                    evidence = driver.find_element("xpath",'.//*[@id="abuse_evidence"]')
                    evidence.send_keys(logs) # This coce  is a generator expression that converts each non-None value in the row list to a string, then joins them together with a space separator.
           

                    # Source IP
                    scr_ip = driver.find_element("xpath",'.//*[@id="source_ip"]')
                    scr_ip.send_keys(row[1])
              

                    # Date
                    date = driver.find_element("xpath",'.//*[@id="date"]')
                    date_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    date_obj = datetime.strptime(date_str, "%b %d, %Y, %I:%M:%S %p") # This code converts a date string in the format of "MMM DD, YYYY, HH:MM:SS AM/PM" into a datetime object in Python.
                    date_input = driver.find_element("xpath", './/*[@id="date"]')
                    date_input.send_keys(date_obj.strftime("%m-%d-%y"))

                    time = driver.find_element("xpath",'.//*[@id="time"]')
                    time_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    time_obj = datetime.strptime(time_str, "%b %d, %Y, %I:%M:%S %p")
                    time_input = driver.find_element("xpath", './/*[@id="time"]')
                    time_input.send_keys(time_obj.strftime("%H:%m"))

                    # Abuse time zone
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-input"]')
                    time_zone.click()
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-option-28"]')
                    time_zone.click()

                    # disktination port/protocol
                    disk_port = row[6]
                    time_input = driver.find_element("xpath", './/*[@id="destination_port"]')
                    time_input.send_keys(disk_port)

                    # destination ip addresses:
                    disk_ip = row[5]
                    time_input = driver.find_element("xpath", './/*[@id="destination_ip_addresses"]')
                    time_input.send_keys(disk_ip)

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()


                    # Submit form
                    submit_report = driver.find_element("xpath",'./html/body/div[1]/div[3]/div/div[1]/section/div/div/div/form/button')
                    submit_report.click()

                    print(logs)

                    row_number += 1
                
                    # Scroll up to the top of the web page
                    driver.execute_script("window.scrollTo(0, 0)")
                    driver.refresh()
                    driver.execute_script("window.scrollTo(0, 0)")
                
                #AbuseTye others        
                else:
                    
                    currRow +=1

                    abtclick = driver.find_element("xpath",'.//*[@id="intrusion_sub_category-others"]')
                    abtclick.click()

                    keyboard.press_and_release('tab')

                    # Name
                    uname = driver.find_element("xpath",'.//*[@id="reporter_name"]')
                    uname.send_keys(name)
                    

                    # Email
                    email_sender = driver.find_element("xpath",'.//*[@id="email"]')
                    email_sender.send_keys(email)
           

                    # Abuse Evidence logs
                    evidence = driver.find_element("xpath",'.//*[@id="abuse_evidence"]')
                    evidence.send_keys(logs) # This coce  is a generator expression that converts each non-None value in the row list to a string, then joins them together with a space separator.
           

                    # Source IP
                    scr_ip = driver.find_element("xpath",'.//*[@id="source_ip"]')
                    scr_ip.send_keys(row[1])
              

                    # Date
                    date = driver.find_element("xpath",'.//*[@id="date"]')
                    date_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    date_obj = datetime.strptime(date_str, "%b %d, %Y, %I:%M:%S %p") # This code converts a date string in the format of "MMM DD, YYYY, HH:MM:SS AM/PM" into a datetime object in Python.
                    date_input = driver.find_element("xpath", './/*[@id="date"]')
                    date_input.send_keys(date_obj.strftime("%m-%d-%y"))

                    time = driver.find_element("xpath",'.//*[@id="time"]')
                    time_str = row[0]  # assuming the date is in the first column of the spreadsheet
                    time_obj = datetime.strptime(time_str, "%b %d, %Y, %I:%M:%S %p")
                    time_input = driver.find_element("xpath", './/*[@id="time"]')
                    time_input.send_keys(time_obj.strftime("%H:%m"))

                    # Abuse time zone
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-input"]')
                    time_zone.click()
                    time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-option-28"]')
                    time_zone.click()

                    # disktination port/protocol
                    disk_port = row[6]
                    time_input = driver.find_element("xpath", './/*[@id="destination_port"]')
                    time_input.send_keys(disk_port)

                    # destination ip addresses:
                    disk_ip = row[5]
                    time_input = driver.find_element("xpath", './/*[@id="destination_ip_addresses"]')
                    time_input.send_keys(disk_ip)

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()

                    #accept agrement
                    agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                    agree.click()


                    # Submit form
                    submit_report = driver.find_element("xpath",'./html/body/div[1]/div[3]/div/div[1]/section/div/div/div/form/button')
                    submit_report.click()

                    print(logs)

                    row_number += 1
                
                    # Scroll up to the top of the web page
                    driver.execute_script("window.scrollTo(0, 0)")
                    driver.refresh()
                    driver.execute_script("window.scrollTo(0, 0)")
            
            #abScantype Others
            else:
                abuseType = row[15]

                driver.refresh()
                driver.refresh()
                driver.refresh()
                
                #ScanType
                atype = driver.find_element(By.ID, 'react-select-3-input')
                atype.click()
                atype = driver.find_element("xpath", './/*[@id="react-select-3-option-2"]')
                atype.click()

                # Name
                uname = driver.find_element("xpath",'.//*[@id="reporter_name"]')
                uname.send_keys(name)
                

                # Email
                email_sender = driver.find_element("xpath",'.//*[@id="email"]')
                email_sender.send_keys(email)
        

                # Abuse Evidence logs
                evidence = driver.find_element("xpath",'.//*[@id="abuse_evidence"]')
                evidence.send_keys(logs) # This coce  is a generator expression that converts each non-None value in the row list to a string, then joins them together with a space separator.
        

                # Source IP
                scr_ip = driver.find_element("xpath",'.//*[@id="source_ip"]')
                scr_ip.send_keys(row[1])
            

                # Date
                date = driver.find_element("xpath",'.//*[@id="date"]')
                date_str = row[0]  # assuming the date is in the first column of the spreadsheet
                date_obj = datetime.strptime(date_str, "%b %d, %Y, %I:%M:%S %p") # This code converts a date string in the format of "MMM DD, YYYY, HH:MM:SS AM/PM" into a datetime object in Python.
                date_input = driver.find_element("xpath", './/*[@id="date"]')
                date_input.send_keys(date_obj.strftime("%m-%d-%y"))

                time = driver.find_element("xpath",'.//*[@id="time"]')
                time_str = row[0]  # assuming the date is in the first column of the spreadsheet
                time_obj = datetime.strptime(time_str, "%b %d, %Y, %I:%M:%S %p")
                time_input = driver.find_element("xpath", './/*[@id="time"]')
                time_input.send_keys(time_obj.strftime("%H:%m"))

                # Abuse time zone
                time_zone = driver.find_element("xpath",'.//*[@id="react-select-4-input"]')
                time_zone.send_keys("be")

                # disktination port/protocol
                disk_port = row[6]
                time_input = driver.find_element("xpath", './/*[@id="destination_port"]')
                time_input.send_keys(disk_port)

                # destination ip addresses:
                disk_ip = row[5]
                time_input = driver.find_element("xpath", './/*[@id="destination_ip_addresses"]')
                time_input.send_keys(disk_ip)

                #accept agrement
                agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                agree.click()

                #accept agrement
                agree = driver.find_element("xpath",'.//*[@id="affirmation_check"]')
                agree.click()


                # Submit form
                submit_report = driver.find_element("xpath",'.//*[@id="__next"]/div[3]/div/div[1]/section/div/div/div/form/button')
                submit_report.click()

                print(logs)

                row_number += 1
            
                # Scroll up to the top of the web page
                driver.execute_script("window.scrollTo(0, 0)")
                driver.refresh()
                driver.execute_script("window.scrollTo(0, 0)")
    else:
        print("Page is not fully loaded")
        driver.quit()

driver.quit()
#Install the chrome driver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
driver.maximize_window()

#Link to be open
driver.get("https://www.youtube.com/@jerhomevlogs5299")
print(f"Excell row # done report: {row_number}")
print("Youtube: Jerhome Vlogs")
time.sleep(10)
driver.quit()
time.sleep(15)
